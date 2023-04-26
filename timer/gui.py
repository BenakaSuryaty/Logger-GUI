import tkinter as tk
from tkinter import ttk
from tkinter import *
import tkinter.messagebox
import tkinter.filedialog
from openpyxl import *
import os
import subprocess
import serial
import time
import serial.tools.list_ports



def validate(P):
    if len(P) == 0:
        return True
    elif len(P) == 1 and P.isdigit():
        return True
    elif len(P) == 2 and P.isdigit():
        return True
    else:
        return False


def ph_val(n):
    if len(n) == 0:
        return True
    elif len(n) == 1 and n.isdigit():
        return True
    elif len(n) == 2 and n.isdigit():
        return True
    elif len(n) == 3 and n.isdigit():
        return True
    elif len(n) == 4 and n.isdigit():
        return True
    elif len(n) == 5 and n.isdigit():
        return True
    elif len(n) == 6 and n.isdigit():
        return True
    elif len(n) == 7 and n.isdigit():
        return True
    elif len(n) == 8 and n.isdigit():
        return True
    elif len(n) == 9 and n.isdigit():
        return True
    elif len(n) == 10 and n.isdigit():
        return True
    else:
        return False


def name_val(a):
    if a.isalpha():
        return True
    elif a == " ":
        return True
    else:
        return False


# Timerapp class
class Timerapp(tk.Tk):
    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)

        self.isConfirmed = False
        self.isSaved = False
        self.isCreated = False
        self.checked = False

       

        self.shared_data = {
            "Name": tk.StringVar(),
            "Age": tk.StringVar(),
            "Phone": tk.StringVar(),
            "Hours": tk.StringVar(),
            "Minutes": tk.StringVar(),
            "Seconds": tk.StringVar(),
        }
        num_keys = ["Hours", "Minutes", "Seconds"]

        for key in num_keys:
            self.shared_data[key].set("00")

        # GUI styling
        tk.Tk.iconbitmap(self, default="stpwtch.ico")
        tk.Tk.title(self, "Timer")
        tk.Tk.geometry(self, "400x300")
        tk.Tk.resizable(self, False, False)

        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {}

        for F in (wlcmpg, prsnlinfo, timerpg):

            frame = F(container, self)

            self.frames[F] = frame

            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame(wlcmpg)

    def show_frame(self, cont):

        frame = self.frames[cont]
        frame.tkraise()

    def pass_text(self, text, text2, text3, text4):
        self.frames[timerpg].Timer(text, text2, text3, text4)

    def pass_filepath(self, name):
        self.frames[prsnlinfo].get_filepath(name)


# Welcome page 
class wlcmpg(tk.Frame):
    def __init__(self, parentcls, controller):
        tk.Frame.__init__(self, parentcls)

        self.controller = controller
        label = tk.Label(self, text="Timer", font=("calibre", 20, "bold"))
        label.place(x=165, y=0)
        frame = tk.LabelFrame(self, text="Genral instructions and steps")
        frame.pack(side="top", fill="x", expand=True, padx=10, pady=10)

        # Assist steps
        sentence = """Step 1: Enter personal details in the \nrespective fields before proceeding.
            \nStep 2: Check details and click "Confirm".
            \nStep 3: The details with time will be saved \t\nin an excel file which can be accessed after the timer 
has elapsed.\n\nNote: In case of a mistake,the data will be available 
         in the excel to edit.
            \n\t    Press "Next" to proceed"""
        Steps = tk.Message(
            frame,
            text=sentence,
            font=("calibre", 10),
            aspect=250,
            anchor="nw",
            justify="left",
        )
        Steps.pack(expand=True)

        button = ttk.Button(self, text="Next >>", command=lambda: pageshift())
        button.place(x=314, y=269)

        button1 = ttk.Button(self, text="Create log file", command=lambda: createfile())
        button1.place(x=9, y=269)

        def createfile():
            self.controller.isCreated = True
            response = tkinter.messagebox.askyesno(
                "Create Log file",
                "Do you want to create a custom excel file in a desired location?",
                icon="question",
            )

            if response:
                files = [("Excel Workbook", "*.xlsx"), ("All Files", "*.*")]
                file = tkinter.filedialog.asksaveasfile(
                    filetypes=files, defaultextension=files
                )
                filepath = file.name
                wb = Workbook()
                wb.save(filepath)
                controller.pass_filepath(str(filepath))

            else:
                desktop = os.path.join(
                    os.path.join(os.environ["USERPROFILE"]), "Desktop"
                )
                filepath = os.path.join(desktop, "Timer_log.xlsx")
                wb = Workbook()
                wb.save(filepath)
                controller.pass_filepath(filepath)
                tkinter.messagebox.showinfo(
                    "File created",
                    " 'Timer_Log.xlsx' is created in your computer's desktop folder",
                )

            button1.destroy()

            button2 = ttk.Button(
                self, text="Open log file", command=lambda: openfile(filepath)
            )
            button2.place(x=9, y=269)

        def openfile(name):
            fiame = os.path.abspath(name)
            cmd = "explorer " + "{}".format(fiame)
            subprocess.call(cmd)

        def pageshift():

            if self.controller.isCreated:
                controller.show_frame(prsnlinfo)
            else:
                tkinter.messagebox.showerror("Log File", "Log file not created!")


# Personal information page
class prsnlinfo(tk.Frame):

    bleh = ""

    def __init__(self, parentcls, controller):
        tk.Frame.__init__(self, parentcls)

        def on_click(event):
            event.widget.delete(0, tk.END)

        self.controller = controller

        frame = tk.LabelFrame(
            self, text="Personal information", font=("calibre", 10, "bold")
        )
        frame.pack(side="left", expand=True, fill="both", padx=10, pady=10, anchor="nw")

        # Entry fields
        self.vcmd = (self.register(validate), "%P")
        self.pcmd = (self.register(ph_val), "%P")
        self.ccmd = (self.register(name_val), "%S")

        self.name_label = tk.Label(frame, text="Name: ", font=("calibre", 10, "bold"))
        self.name_entry = tk.Entry(
            frame,
            validate="key",
            validatecommand=self.ccmd,
            textvariable=controller.shared_data["Name"],
            font=("calibre", 10, "normal"),
            border=3,
            borderwidth=2,
        )
        self.name_entry.bind("<Button>", on_click)

        self.age_label = tk.Label(frame, text="Age: ", font=("calibre", 10, "bold"))
        self.age_entry = tk.Entry(
            frame,
            validate="key",
            validatecommand=self.vcmd,
            textvariable=controller.shared_data["Age"],
            font=("calibre", 10, "normal"),
            border=3,
            borderwidth=2,
        )
        self.age_entry.bind("<Button>", on_click)

        self.phone_label = tk.Label(
            frame, text="Phone Number: ", font=("calibre", 10, "bold")
        )
        self.phone_entry = tk.Entry(
            frame,
            validate="key",
            validatecommand=self.pcmd,
            textvariable=controller.shared_data["Phone"],
            font=("calibre", 10, "normal"),
            border=3,
            borderwidth=2,
        )
        self.phone_entry.bind("<Button>", on_click)

        self.time_label = tk.Label(
            frame, text="Time Limit: ", font=("calibre", 10, "bold")
        )

        self.hour_lable = tk.Label(frame, text="Hr: ", font=("calibre", 10, "bold"))
        self.hour_entry = tk.Entry(
            frame,
            validate="key",
            validatecommand=self.vcmd,
            textvariable=controller.shared_data["Hours"],
            font=("calibre", 10, "normal"),
            border=3,
            borderwidth=2,
        )
        self.hour_entry.bind("<Button>", on_click)

        self.min_lable = tk.Label(frame, text="Min: ", font=("calibre", 10, "bold"))
        self.min_entry = tk.Entry(
            frame,
            validate="key",
            validatecommand=self.vcmd,
            textvariable=controller.shared_data["Minutes"],
            font=("calibre", 10, "normal"),
            border=3,
            borderwidth=2,
        )
        self.min_entry.bind("<Button>", on_click)

        self.sec_lable = tk.Label(frame, text="Sec: ", font=("calibre", 10, "bold"))
        self.sec_entry = tk.Entry(
            frame,
            validate="key",
            validatecommand=self.vcmd,
            textvariable=controller.shared_data["Seconds"],
            font=("calibre", 10, "normal"),
            border=3,
            borderwidth=2,
        )
        self.sec_entry.bind("<Button>", on_click)

        self.button1 = ttk.Button(self, text="<< Back", command=lambda: warnreturn())
        self.button2 = ttk.Button(self, text="Next >>", command=lambda: checksave())
        self.button3 = ttk.Button(self, text="Confirm", command=lambda: checkempty())
        self.button4 = ttk.Button(
            self, text="Save to file", command=lambda: self.submit()
        )

        self.name_label.place(x=10, y=10)
        self.name_entry.place(x=60, y=10, width=300)
        self.age_label.place(x=10, y=50)
        self.age_entry.place(x=50, y=50, width=20)
        self.phone_label.place(x=10, y=90)
        self.phone_entry.place(x=120, y=90)
        self.time_label.place(x=10, y=125)
        self.hour_lable.place(x=90, y=145)
        self.hour_entry.place(x=115, y=145, width=20)
        self.min_lable.place(x=150, y=145)
        self.min_entry.place(x=185, y=145, width=20)
        self.sec_lable.place(x=210, y=145)
        self.sec_entry.place(x=245, y=145, width=20)

        self.button1.place(x=9, y=264)
        self.button2.place(x=314, y=264)
        self.button3.place(x=114, y=264)
        self.button4.place(x=214, y=264)

        def send_text(text, text2, text3, text4):

            self.controller.isConfirmed = True
            controller.pass_text(text, text2, text3, text4)
            self.popup()

        def checkempty():

            if self.name_entry.get() == "":
                tkinter.messagebox.showerror("Error", "Enter Name!")
            elif self.age_entry.get() == "":
                tkinter.messagebox.showerror("Error", "Enter Age!")
            elif self.phone_entry.get() == "":
                tkinter.messagebox.showerror("Error", "Enter Phone number!")
            elif len(self.phone_entry.get()) < 10:
                tkinter.messagebox.showerror("Error", "Incorrect Phone number!")
            elif (
                self.hour_entry.get()
                == self.min_entry.get()
                == self.sec_entry.get()
                == "00"
                or self.hour_entry.get()
                == self.min_entry.get()
                == self.sec_entry.get()
                == ""
            ):
                tkinter.messagebox.showerror("Error", "Inappropriate Time!")
            else:
                send_text(
                    controller.shared_data["Name"].get(),
                    controller.shared_data["Hours"].get(),
                    controller.shared_data["Minutes"].get(),
                    controller.shared_data["Seconds"].get(),
                )

        def checksave():

            if controller.isSaved:
                controller.show_frame(timerpg)
            else:
                tkinter.messagebox.showwarning(
                    "Warning!",
                    "Save the entered data Before proceeding!"
                    + '\nClick "Save to file" to Save.',
                )

        def warnreturn():

            response = tkinter.messagebox.askyesno(
                "Warning!",
                "Are you sure to exit?\nProceeding to exit would erase the entered data.",
                icon="warning",
            )
            if response:
                self.name_entry.delete(0, tk.END)
                self.age_entry.delete(0, tk.END)
                self.phone_entry.delete(0, tk.END)
                self.hour_entry.delete(0, tk.END)
                self.min_entry.delete(0, tk.END)
                self.sec_entry.delete(0, tk.END)
                self.deletedata()
                self.button4["state"] = "Enable"
                num_keys = ["Hours", "Minutes", "Seconds"]
                for key in num_keys:
                    controller.shared_data[key].set("00")
                misc_keys = ["Phone", "Age"]
                for key in misc_keys:
                    controller.shared_data[key].set(0)
                controller.show_frame(wlcmpg)

    def get_filepath(self, path):
        self.bleh = path

    def submit(self):

        if self.controller.isConfirmed:
            book = load_workbook(self.bleh)
            sheet = book.active

            sheet["A1"] = "Name"
            sheet["B1"] = "Age"
            sheet["C1"] = "Phone"
            sheet["D1"] = "Hours"
            sheet["E1"] = "Minutes"
            sheet["F1"] = "Seconds"

            msg = "Successfully saved to {}!".format(os.path.basename(self.bleh))
            adata = {}
            Keys = ["Age", "Phone", "Hours", "Minutes", "Seconds"]

            maxrow = sheet.max_row

            adata["Name"] = self.controller.shared_data["Name"].get()
            for key in Keys:
                if self.controller.shared_data[key].get() == "":
                    self.controller.shared_data[key].set("00")
                    adata[key] = int(self.controller.shared_data[key].get())

                else:
                    adata[key] = int(self.controller.shared_data[key].get())

            for column, value in enumerate(adata.values(), start=1):
                sheet.cell(row=(maxrow + 1), column=column).value = value

            book.save(self.bleh)

            self.controller.isSaved = True
            tkinter.messagebox.showinfo("Saved", msg)
            self.button4["state"] = "disabled"

        else:
            tkinter.messagebox.showwarning(
                "Warning!",
                "Check the entered data Before proceeding!"
                + '\nClick "Confirm" to check.',
            )

    def deletedata(self):
        book = load_workbook(self.bleh)
        sheet = book.active
        maxrow = sheet.max_row
        sheet.delete_rows(maxrow)
        book.save(self.bleh)

    def popup(self):

        Name = self.controller.shared_data["Name"].get()
        Age = self.controller.shared_data["Age"].get()
        Phone = self.controller.shared_data["Phone"].get()
        Hours = self.controller.shared_data["Hours"].get()
        Minutes = self.controller.shared_data["Minutes"].get()
        Seconds = self.controller.shared_data["Seconds"].get()

        lst = [Hours, Minutes, Seconds]

        for i in range(len(lst)):
            if lst[i] == "":
                lst[i] = "0"

        prnt = """Do you wish to edit the entered data?\n Name: {}\n Age: {}\n Phone Number: {}\n Time:\n     Hours: {}  Minutes: {}  Seconds: {}
        """.format(
            Name, Age, Phone, lst[0], lst[1], lst[2]
        )
        tkinter.messagebox.askquestion("Verify", message=prnt)

def find_com():
    lsp = []
    ports = list(serial.tools.list_ports.comports())
    for p in ports:
        if "CH340" in p.description:
            lsp = str(p).split()
    if len(lsp) !=0:
        Com = lsp[0]
        return Com
    else:
        raise Exception("Check the Board's connection...")

com = find_com()
            
arduino = serial.Serial(port= com, baudrate=115200, timeout=.1)
# Timer data display page

class timerpg(tk.Frame):

    ls = []
    com = "COM3"
    stop = True
    
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.controller = controller

    def ispresent(self):
        self.controller.checked = True
        if com != "":
            tkinter.messagebox.showinfo("Timer Board Found", "The board is connected.")
        else:
            tkinter.messagebox.showwarning(
                "Timer Board Not found", "Please check the board connection."
            )
    
    
    def stp(self):
        if arduino.readline().decode():
            return False
        else:
            return True
        
    def sendstp(self):
        
        self.stop = False
        
    
    def send_data(self,x):
        
        arduino.write(bytes(x, "utf-8"))
        arduino.flushOutput()
    
    
    def Timer(self, Name, Hours, Minutes, Seconds):

        """
        Remember the indices:
        ls[0] = Hours
        ls[1] = Minutes
        ls[2] = Seconds
        """
        self.ls = [Hours, Minutes, Seconds]

        for i in range(len(self.ls)):
            if self.ls[i] == "":
                self.ls[i] = "0"
        
        Hours = self.ls[0]
        Minutes = self.ls[1]
        Seconds = self.ls[2]
        
        text = "{}'s remaining time:".format(Name)
        self.dispstatusmsg = tk.Label(self, text=text, font=("calibre", 14),
                                      bg= "black",fg= "turquoise").place(x=20,y=40)
        
        self.starticon = PhotoImage(file=r"D:/Do not touch/Timerprg/timer/play-button.png")
        self.startimage = self.starticon.subsample(2, 2)
        self.stopicon = PhotoImage(file=r"D:/Do not touch/Timerprg/timer/stop-button.png")
        self.stopimage = self.stopicon.subsample(2, 2)
        
        self.button1 = ttk.Button(
            self,
            text="Start Timer",
            image=self.startimage,
            compound=LEFT,
            command=lambda: start(),
        ).place(x=10, y=255)
        
        self.button2 = ttk.Button(
            self, text="Stop Timer", 
            image=self.stopimage, 
            compound=LEFT,
            command=lambda: self.sendstp()).place(x=285, y=255)
        
        self.button3 = ttk.Button(
            self, text="Check for Timer connection", command=lambda: self.ispresent()
        ).place(x=122, y=225)

        def start():
            msg = tk.Message(self, font=("calibre", 58), width=300, anchor="w")
            times = int(Hours) * 3600 + int(Minutes) * 60 + int(Seconds)
                
            while times > -1:
                if self.stop:
                    minute, second = (times // 60, times % 60)
                    hour = 0
                    if minute > 60:
                        hour, minute = (minute // 60, minute % 60)
                
                    byt = "{}:{}:{}".format(hour,minute,second) 
                    msg.config(text=byt)
                    msg.place(x=40, y=80)
                
                    sndbyt = "{}-{}:{}:{}-".format(Name,hour,minute,second)

                    self.send_data(sndbyt)
                    self.update()
                    
                    time.sleep(1)
                    if times == 0:
                        msg.config(text="0:0:0")
                        msg.place(x=40, y=80)
                    times -= 1
                
                else:
                    msg.destroy()
                    self.dispstatusmsg = tk.Label(self, text="Timer Stopped!", font=("calibre", 30),
                                      bg= "black",fg= "turquoise").place(x=60, y=140)
                    stpbit = "0"
                    arduino.write(bytes(stpbit, 'utf-8'))
                    break
            
            

# Driver code
if __name__ == "__main__":
    app = Timerapp()

    def on_closing():
        if tkinter.messagebox.askokcancel("Quit", "Do you want to quit?"):
            app.destroy()

    app.protocol("WM_DELETE_WINDOW", on_closing)
    app.mainloop()